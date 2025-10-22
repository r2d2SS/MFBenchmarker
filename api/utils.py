import pandas as pd
import io
from decimal import Decimal
from django.db.models import Q
from .models import Product, ScrapedData, MarginSetting, ExchangeRate


def calculate_benchmark_prices(stock_filter='instock', brand_filter=None):
    """
    Calculate benchmark prices for products based on competitors and margins
    """
    # Get current margin and exchange rate
    margin_percent = MarginSetting.get_current_margin()
    margin_multiplier = Decimal(str(1 - (margin_percent / 100)))
    exchange_rate = Decimal(str(ExchangeRate.get_current_rate()))
    
    # Base queryset
    products = Product.objects.select_related('brand', 'supplier')
    
    if stock_filter != 'all':
        products = products.filter(stock_status=stock_filter)
    
    if brand_filter:
        products = products.filter(brand__name__icontains=brand_filter)
    
    benchmark_data = []
    
    for product in products:
        # Get latest scraped data for this product
        scraped_data = ScrapedData.objects.filter(
            Q(product=product) | Q(product_name__icontains=product.name[:20])
        ).select_related('competitor')
        
        # Initialize price dict
        price_data = {
            'sku': product.sku,
            'brand': product.brand.name,
            'product_name': product.name,
            'stock_status': product.stock_status,
            # Supplier prices (calculated with margin)
            'ta_price': None, 'co_price': None, 'dv_price': None,
            'ge_price': None, 'lp_price': None, 'ra_price': None,
            'la_price': None, 'cf_price': None,
            # Competitor prices
            'ih_price': None, 'ch_price': None, 'nt_price': None,
        }
        
        # Calculate supplier prices if cost is available
        if product.cost_with_tax and product.supplier:
            supplier_price = (product.cost_with_tax / margin_multiplier) * exchange_rate
            supplier_code = product.supplier.code.lower()
            if f"{supplier_code}_price" in price_data:
                price_data[f"{supplier_code}_price"] = round(supplier_price, 2)
        
        # Add competitor prices
        competitor_prices = []
        for data in scraped_data:
            competitor_code = data.competitor.code.lower()
            if f"{competitor_code}_price" in price_data:
                price_data[f"{competitor_code}_price"] = float(data.price)
                competitor_prices.append(float(data.price))
        
        # Calculate competitor statistics
        if competitor_prices:
            price_data['min_competitor_price'] = min(competitor_prices)
            price_data['max_competitor_price'] = max(competitor_prices)
            price_data['avg_competitor_price'] = sum(competitor_prices) / len(competitor_prices)
            price_data['recommended_price'] = price_data['avg_competitor_price'] * 0.95  # 5% below avg
        else:
            price_data.update({
                'min_competitor_price': None,
                'max_competitor_price': None, 
                'avg_competitor_price': None,
                'recommended_price': None
            })
        
        benchmark_data.append(price_data)
    
    return benchmark_data


def generate_excel_report(products, report_type='instock'):
    """
    Generate Excel report with formatting
    """
    # Get benchmark data
    benchmark_data = calculate_benchmark_prices(
        stock_filter=report_type if report_type != 'never_sold' else 'all'
    )
    
    # Convert to DataFrame
    df = pd.DataFrame(benchmark_data)
    
    if df.empty:
        df = pd.DataFrame([{'message': 'No data available'}])
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=report_type.title(), index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[report_type.title()]
        
        # Add formatting
        from openpyxl.styles import PatternFill, Font
        from openpyxl.formatting.rule import CellIsRule
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Conditional formatting for prices
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Apply conditional formatting to price columns (assuming columns D onwards are prices)
        if len(df.columns) > 4:
            price_range = f"D2:Z{len(df) + 1}"
            worksheet.conditional_formatting.add(
                price_range,
                CellIsRule(operator="greaterThan", formula=["AVERAGE($D$2:$Z$2)"], fill=red_fill)
            )
            worksheet.conditional_formatting.add(
                price_range,
                CellIsRule(operator="lessThan", formula=["AVERAGE($D$2:$Z$2)"], fill=green_fill)
            )
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output


def format_currency(value, currency='USD'):
    """Format decimal value as currency"""
    if value is None:
        return None
    return f"{currency} {value:.2f}"


def calculate_margin_impact(old_margin, new_margin, products=None):
    """
    Calculate the impact of margin changes on pricing
    """
    if products is None:
        products = Product.objects.filter(cost_with_tax__isnull=False)
    
    old_multiplier = Decimal(str(1 - (old_margin / 100)))
    new_multiplier = Decimal(str(1 - (new_margin / 100)))
    
    impact_data = []
    
    for product in products:
        if product.cost_with_tax:
            old_price = product.cost_with_tax / old_multiplier
            new_price = product.cost_with_tax / new_multiplier
            price_change = new_price - old_price
            percent_change = (price_change / old_price) * 100
            
            impact_data.append({
                'sku': product.sku,
                'product_name': product.name,
                'old_price': float(old_price),
                'new_price': float(new_price),
                'price_change': float(price_change),
                'percent_change': float(percent_change)
            })
    
    return impact_data


def update_exchange_rates():
    """
    Update exchange rates from external API
    TODO: Implement with real currency API
    """
    from datetime import date
    from .models import ExchangeRate
    
    # Placeholder - would integrate with real currency API
    try:
        # Example rate - replace with actual API call
        rate_data = {
            'EUR_USD': 1.0533,
            'USD_EUR': 0.94925
        }
        
        for currency_pair, rate in rate_data.items():
            from_curr, to_curr = currency_pair.split('_')
            
            ExchangeRate.objects.update_or_create(
                from_currency=from_curr,
                to_currency=to_curr,
                date=date.today(),
                defaults={'rate': Decimal(str(rate))}
            )
        
        return True
    except Exception as e:
        print(f"Error updating exchange rates: {e}")
        return False